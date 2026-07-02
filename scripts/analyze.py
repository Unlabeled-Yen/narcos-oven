# -*- coding: utf-8 -*-
"""narcos-oven 資料完整分析器 v4
- atom / product id 全中文
- 從三個 xlsx 歸納：品項、組合、通路自由文字、KOL 合成邊界、異常
- ⚠️ 不再覆寫 data/menu.yaml（Yen 手改為主）
- 只輸出 docs/data-analysis.md（含「建議 menu.yaml 修改」一節）

用法：
    python3 scripts/analyze.py [fixture_dir]
預設 fixture_dir = fixtures/2026-07-round1
"""
import openpyxl, os, re, sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
FIXTURE = sys.argv[1] if len(sys.argv) > 1 else str(ROOT / "fixtures/2026-07-round1")
OUT_MD = ROOT / "docs/data-analysis.md"
OUT_YAML = ROOT / "data/menu.yaml"

def to_num(v):
    """Excel 常把數字存成字串，安全轉。"""
    if v is None: return None
    if isinstance(v, (int, float)): return v
    try: return float(str(v).replace(",", "").strip())
    except: return None

def ws(f, sheet_name):
    wb = openpyxl.load_workbook(os.path.join(FIXTURE, f), data_only=True)
    return wb[sheet_name]

# ============================================================
# 1. 賣貨便
# ============================================================
seller_buy_orders = []
seller_buy_products = Counter()
seller_buy_c12_shipping = []
seller_buy_c28_notes = []
c22_over1 = []

for f in ["1.xlsx", "2.xlsx"]:
    sh = ws(f, "非訂單匯入")
    current = None
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i < 3: continue
        r = list(row)
        if len(r) < 29: r = r + [None] * (29 - len(r))
        c4, c5, c12, c13, c15, c16, c17, c18, c19, c20, c21, c22, c28 = (
            r[4], r[5], r[12], r[13], r[15], r[16], r[17],
            r[18], r[19], r[20], r[21], r[22], r[28]
        )
        if c4:
            if current: seller_buy_orders.append(current)
            current = {
                "order_id": c4, "src_file": f,
                "status": str(c5) if c5 else "", "items": [],
                "c22": to_num(c22), "freight": to_num(c17), "notes": c28,
                "total": to_num(c21),
                "d_c18": to_num(c18) or 0, "d_c19": to_num(c19) or 0, "d_c20": to_num(c20) or 0,
            }
            if c12:
                current["items"].append({"name": str(c12).strip(), "price": to_num(c13), "qty": to_num(c15), "subtotal": to_num(c16)})
                seller_buy_products[str(c12).strip()] += 1
            if c28: seller_buy_c28_notes.append(str(c28).strip())
            if current["c22"] is not None and current["c22"] > 1:
                c22_over1.append({"id": c4, "c22": int(current["c22"]), "file": f})
        else:
            if c12 and current:
                s = str(c12).strip()
                current["items"].append({"name": s, "price": to_num(c13), "qty": to_num(c15), "subtotal": to_num(c16)})
                seller_buy_products[s] += 1
        if c12 and "指定出貨日" in str(c12):
            seller_buy_c12_shipping.append(str(c12).strip())
    if current: seller_buy_orders.append(current)

# ============================================================
# 2. 面交
# ============================================================
in_person_c2 = []
in_person_products = Counter()
in_person_product_headers = []
sh = ws("2026 六月 面交訂購單 (回覆).xlsx", "表單回覆 1")
header = list(sh.iter_rows(min_row=1, max_row=1, values_only=True))[0]
in_person_product_headers = [(idx, header[idx]) for idx in range(7, 21) if header[idx]]
for i, row in enumerate(sh.iter_rows(values_only=True)):
    if i == 0: continue
    r = list(row)
    if len(r) < 29: r = r + [None] * (29 - len(r))
    c2 = r[2]
    if c2 and str(c2).strip():
        in_person_c2.append((i+1, str(c2).strip()))
    for col_idx in range(7, 21):
        val = to_num(r[col_idx]) if col_idx < len(r) else None
        if val and val > 0:
            in_person_products[header[col_idx]] += int(val)

# ============================================================
# 3. KOL
# ============================================================
sh = ws("KOL 合作.xlsx", "未完成")
kol_records = []
kol_all_products = Counter()
kol_c4_shipping = Counter()
current_kol = None
for i, row in enumerate(sh.iter_rows(values_only=True)):
    if i == 0: continue
    r = list(row)
    if len(r) < 11: r = r + [None] * (11 - len(r))
    c1, c2, c3, c4, c5, c6 = r[1], r[2], r[3], r[4], r[5], r[6]
    has_c1 = c1 and str(c1).strip()
    has_c2 = c2 and str(c2).strip()
    if has_c1 or has_c2:
        if current_kol: kol_records.append(current_kol)
        current_kol = {
            "row_start": i+1,
            "discount_code": str(c1).strip() if c1 else None,
            "ig": str(c2).strip() if c2 else None,
            "followers": str(c3).strip() if c3 else None,
            "ship_date_raw": c4,
            "products": [],
            "shipped": bool(c6) if c6 is not None else None,
        }
        if c5: current_kol["products"].append(str(c5).strip())
        if c4:
            kol_c4_shipping[str(c4)[:30]] += 1
    else:
        if c5 and current_kol:
            current_kol["products"].append(str(c5).strip())
if current_kol: kol_records.append(current_kol)
for k in kol_records:
    for p in k["products"]:
        kol_all_products[p] += 1

# ============================================================
# 4. 分析
# ============================================================
shipping_date_re = re.compile(r'指定出貨日.*?(\d+)/(\d+)（(.)）', re.UNICODE)
shipping_dates_extracted = []
shipping_dates_failed = []
for s in seller_buy_c12_shipping:
    m = shipping_date_re.search(s)
    if m: shipping_dates_extracted.append({"raw": s[:80], "month": m.group(1), "day": m.group(2), "weekday": m.group(3)})
    else: shipping_dates_failed.append(s[:80])

c2_channel_re = re.compile(r'(中壢|台中|台北|新竹|高雄|桃園|台南|嘉義|澎湖)')
c2_date_re = re.compile(r'(\d+)/(\d+)')
c2_type_re = re.compile(r'(面交|冷凍宅配|宅配|駐店|活動|私定)')
c2_analysis = []
c2_full = 0; c2_partial = 0; c2_none = 0
for row_num, val in in_person_c2:
    chan_m = c2_channel_re.search(val)
    date_m = c2_date_re.search(val)
    type_m = c2_type_re.search(val)
    r = {"row": row_num, "raw": val,
         "location": chan_m.group(1) if chan_m else None,
         "date": f"{date_m.group(1)}/{date_m.group(2)}" if date_m else None,
         "type": type_m.group(1) if type_m else None}
    c2_analysis.append(r)
    hits = sum(1 for x in (r["location"], r["date"], r["type"]) if x)
    if hits == 3: c2_full += 1
    elif hits > 0: c2_partial += 1
    else: c2_none += 1

seller_buy_mismatches = []
seller_buy_ok = 0
for o in seller_buy_orders:
    if not o["items"] or o["total"] is None: continue
    sub_sum = sum((it["subtotal"] or 0) for it in o["items"] if it["subtotal"] is not None)
    expected = sub_sum + (o["freight"] or 0) - o["d_c18"] - o["d_c19"] - o["d_c20"]
    diff = abs(expected - o["total"])
    if diff > 2:
        seller_buy_mismatches.append({
            "id": o["order_id"], "sub_sum": sub_sum, "freight": o["freight"] or 0,
            "discount": o["d_c18"] + o["d_c19"] + o["d_c20"],
            "expected": expected, "total": o["total"], "diff": diff})
    else:
        seller_buy_ok += 1

# ============================================================
# 5. menu.yaml v0 (全中文 identifier)
# ============================================================
ATOMS = {
    "肉桂捲":              {"unit": "顆"},
    "蘋果肉桂捲":          {"unit": "顆"},
    "焦糖蘋果肉桂麵包":    {"unit": "顆"},
    "芝麻焙茶奶酥磅蛋糕":  {"unit": "顆"},
    "鳳梨肉桂奶酥磅蛋糕":  {"unit": "顆"},
    "原味巴斯克":          {"unit": "顆"},
    "白玉原味巴斯克":      {"unit": "顆"},
    "芝麻巴斯克":          {"unit": "顆"},
    "白玉芝麻巴斯克":      {"unit": "顆"},
    "焙茶栗子巴斯克":      {"unit": "顆"},
    "焙茶巴斯克":          {"unit": "顆"},   # ⚠️ 面交表獨有、待雇主 confirm 是否為獨立品項
    "白玉烏龍茶巴斯克":    {"unit": "顆"},
    "香料堅果醬40ml":      {"unit": "罐"},
    "香料堅果醬90ml":      {"unit": "罐"},
    "香料堅果醬240ml":     {"unit": "罐"},
    "瑕疵小脆捲10顆":      {"unit": "包"},
}

def guess_sku(name):
    """回傳 (sku_id, category, atoms:[(atom_zh, count)], price)"""
    n = name
    if "四顆蘋果肉桂捲" in n and "禮盒" in n:
        return ("蘋果肉桂捲4入", "combo", [("蘋果肉桂捲", 4)], 480)
    if "四顆肉桂捲" in n and "禮盒" in n:
        return ("經典肉桂捲4入", "combo", [("肉桂捲", 4)], 400)
    if "五顆蘋果肉桂捲＋一盒香料堅果醬" in n:
        return ("蘋果肉桂捲5入含醬", "combo", [("蘋果肉桂捲", 5), ("香料堅果醬90ml", 1)], 700)
    if "五顆肉桂捲＋一盒香料堅果醬" in n:
        return ("經典肉桂捲5入含醬", "combo", [("肉桂捲", 5), ("香料堅果醬90ml", 1)], 600)
    if "三顆肉桂捲＋兩顆蘋果肉桂捲＋一個香料堅果醬" in n:
        return ("混合5入含醬", "combo", [("肉桂捲", 3), ("蘋果肉桂捲", 2), ("香料堅果醬90ml", 1)], 640)
    if "兩顆肉桂捲＋一顆焦糖蘋果肉桂麵包" in n:
        return ("方型3入_焦糖蘋果", "combo", [("肉桂捲", 2), ("焦糖蘋果肉桂麵包", 1)], 400)
    if "兩顆肉桂捲＋一顆芝麻焙茶奶酥磅蛋糕" in n:
        return ("方型3入_芝麻磅", "combo", [("肉桂捲", 2), ("芝麻焙茶奶酥磅蛋糕", 1)], 560)
    if "兩顆肉桂捲＋一顆鳳梨肉桂奶酥磅蛋糕" in n:
        return ("方型3入_鳳梨磅", "combo", [("肉桂捲", 2), ("鳳梨肉桂奶酥磅蛋糕", 1)], 560)
    if "兩顆肉桂捲＋芝麻焙茶奶酥＋鳳梨肉桂磅蛋糕" in n:
        return ("長型4入_兩磅", "combo", [("肉桂捲", 2), ("芝麻焙茶奶酥磅蛋糕", 1), ("鳳梨肉桂奶酥磅蛋糕", 1)], 920)
    if "蘋果肉桂麵包＋芝麻焙茶奶酥＋鳳梨肉桂磅蛋糕" in n:
        return ("長型3入_三磅", "combo", [("焦糖蘋果肉桂麵包", 1), ("芝麻焙茶奶酥磅蛋糕", 1), ("鳳梨肉桂奶酥磅蛋糕", 1)], 920)
    if "香料堅果醬" in n:
        if "90" in n: return ("香料堅果醬90ml", "single", [("香料堅果醬90ml", 1)], 200)
        if "240" in n: return ("香料堅果醬240ml", "single", [("香料堅果醬240ml", 1)], 500)
        if "40" in n: return ("香料堅果醬40ml", "single", [("香料堅果醬40ml", 1)], None)
    if "白玉烏龍茶巴斯克" in n or "白玉烏龍" in n:
        return ("白玉烏龍茶巴斯克", "single", [("白玉烏龍茶巴斯克", 1)], 980)
    if "白玉原味巴斯克" in n:
        return ("白玉原味巴斯克", "single", [("白玉原味巴斯克", 1)], 750)
    if "白玉芝麻巴斯克" in n:
        return ("白玉芝麻巴斯克", "single", [("白玉芝麻巴斯克", 1)], None)
    if "焙茶栗子巴斯克" in n or "焙茶巴斯克" in n:
        return ("焙茶栗子巴斯克", "single", [("焙茶栗子巴斯克", 1)], 880)
    if "芝麻巴斯克" in n:
        return ("芝麻巴斯克", "single", [("芝麻巴斯克", 1)], 720)
    if "原味巴斯克" in n:
        return ("原味巴斯克", "single", [("原味巴斯克", 1)], 650)
    if "瑕疵小脆捲" in n:
        return ("瑕疵小脆捲10顆", "single", [("瑕疵小脆捲10顆", 1)], 500)
    return None

menu = {"products": {}}
unrecognized = []
all_seller = {k: v for k, v in seller_buy_products.items() if "指定出貨日" not in k}
for name, count in sorted(all_seller.items(), key=lambda x: -x[1]):
    g = guess_sku(name)
    if not g:
        unrecognized.append({"name": name, "count": count})
        continue
    sku_id, category, atoms, price = g
    if sku_id not in menu["products"]:
        menu["products"][sku_id] = {
            "display_name": name, "category": category, "aliases": [],
            "contains": [{"atom": a, "count": c} for a, c in atoms],
            "price": price, "cost": None, "seen_count": count,
        }
    else:
        if name != menu["products"][sku_id]["display_name"]:
            menu["products"][sku_id]["aliases"].append(name)
        menu["products"][sku_id]["seen_count"] += count

def guess_meet_to_sku(meet_name):
    if "肉桂捲四入" in meet_name and "蘋果" not in meet_name: return "經典肉桂捲4入"
    if "蘋果肉桂捲四入" in meet_name: return "蘋果肉桂捲4入"
    if "焦糖蘋果肉桂麵包" in meet_name: return "方型3入_焦糖蘋果"
    if "芝麻焙茶奶酥磅蛋糕" in meet_name: return "方型3入_芝麻磅"
    if "鳳梨肉桂奶酥磅蛋糕" in meet_name: return "方型3入_鳳梨磅"
    if "肉桂捲x5" in meet_name and "香料堅果醬" in meet_name: return "經典肉桂捲5入含醬"
    if "蘋果肉桂捲x5" in meet_name: return "蘋果肉桂捲5入含醬"
    if "肉桂捲x3" in meet_name and "蘋果肉桂捲x2" in meet_name: return "混合5入含醬"
    if "原味巴斯克" in meet_name: return "原味巴斯克"
    if "芝麻巴斯克" in meet_name: return "芝麻巴斯克"
    if "焙茶巴斯克" in meet_name: return "⚠️需雇主 confirm（焙茶巴斯克 vs 焙茶栗子巴斯克）"
    if "白玉烏龍" in meet_name: return "白玉烏龍茶巴斯克"
    if "90ml" in meet_name and "香料堅果醬" in meet_name: return "香料堅果醬90ml"
    if "240ml" in meet_name and "香料堅果醬" in meet_name: return "香料堅果醬240ml"
    return None

# ============================================================
# 6. 輸出 markdown
# ============================================================
batch_dist = Counter()
for s in shipping_dates_extracted:
    batch_dist[f"{s['month']}/{s['day']}"] += 1

md = []
md.append(f"""# narcos-oven 資料分析 v0

_資料來源：`fixtures/2026-07-round1/`_
_處理原則：憲章 #1（靜默失效零容忍）+ #2（AI native、LLM 放對位置）_
_識別碼慣例：全中文（atoms + products id）_

## 📌 TL;DR

| 資料源 | 規模 |
|---|---|
| **賣貨便**（1.xlsx + 2.xlsx，非訂單匯入 sheet） | **{len(seller_buy_orders)} 筆訂單**、{sum(len(o['items']) for o in seller_buy_orders)} 個品項行 |
| **面交**（2026 六月 面交訂購單） | **{len(in_person_c2)} 筆有效面交單** |
| **KOL**（未完成 sheet） | **{len(kol_records)} 位 KOL 合作紀錄** |

## 💥 憲章實例（本次分析發現的靜默失效）

**bug**：初版 `isinstance(c22, (int, float))` 檢查未通過——Excel 常把數字存成字串 `'2'` 而非 int。結果 c22>1 訂單本應有 **{len(c22_over1)} 筆**，初版靜默返回 **0 筆**、產出報表寫「0 筆」。

修正：加 `to_num()` helper 統一處理 str/int/float。

**啟示**：Domain code 每個 numeric 欄位都要走 `to_num`，不能信 Excel 的資料型態。

---

## 1. 品項/組合完整歸納（menu.yaml v0）

### 1.1 賣貨便所有 c12「商品名稱」

共 **{len(all_seller)} 個 unique 品項字串**（去掉「指定出貨日」marker）。

| # | 商品名（原始） | 次數 | SKU | 原子拆解 |
|---|---|---|---|---|""")
for i, (name, count) in enumerate(sorted(all_seller.items(), key=lambda x: -x[1])):
    g = guess_sku(name)
    if g:
        sku_id, cat, atoms, price = g
        atoms_str = " + ".join(f"{a}×{c}" for a,c in atoms)
        md.append(f"| {i+1} | `{name[:60]}` | {count} | `{sku_id}` ({cat}) | {atoms_str} |")
    else:
        md.append(f"| {i+1} | `{name[:60]}` | {count} | ❌ **未歸類** | 需 Yen 補 |")

md.append(f"""

### 1.2 未歸類品項

{'✅ 全部歸類完成！' if not unrecognized else '⚠️ 以下需你確認：'}""")
for u in unrecognized:
    md.append(f"- `{u['name']}`（出現 {u['count']} 次）")

md.append(f"""

### 1.3 面交表品項 header（c7-c20）與賣貨便對照

| 欄 | 面交欄名 | 本次收單 | 建議對應 SKU |
|---|---|---|---|""")
for col_idx, name in in_person_product_headers:
    count = in_person_products.get(name, 0)
    guess = guess_meet_to_sku(name)
    md.append(f"| c{col_idx} | {name} | {count} | `{guess}` |" if guess else f"| c{col_idx} | {name} | {count} | ⚠️ 未對到 |")

md.append(f"""

### 1.4 KOL 「提供項目」欄的品項

未完成 sheet 共 {len(kol_records)} 位 KOL，共 **{len(kol_all_products)} 個 unique 品項字串**：

| # | KOL 品項字串 | 出現次數 |
|---|---|---|""")
for i, (p, c) in enumerate(sorted(kol_all_products.items(), key=lambda x: -x[1])[:30]):
    md.append(f"| {i+1} | `{p}` | {c} |")

md.append(f"""

⚠️ **KOL 品項命名比賣貨便更口語**（例：「四入肉桂捲」vs「經典肉桂捲 四入禮盒」）。menu.yaml 要建 alias mapping。

---

## 2. 面交 c2「在哪邊取貨！」全樣本 + regex 覆蓋率

### 2.1 覆蓋率

| 狀態 | 筆數 | 比例 |
|---|---|---|
| ✅ 全解析（通路+日期+類型都抓到） | {c2_full} | {100*c2_full/max(len(in_person_c2),1):.1f}% |
| 🟡 部分解析 | {c2_partial} | {100*c2_partial/max(len(in_person_c2),1):.1f}% |
| 🔴 完全打不下 → 待處理桶 | {c2_none} | {100*c2_none/max(len(in_person_c2),1):.1f}% |

### 2.2 建議 regex

```python
LOCATION_RE = r'(中壢|台中|台北|新竹|高雄|桃園|台南|嘉義|澎湖)'
DATE_RE     = r'(\\d+)/(\\d+)'
TIME_RE     = r'(\\d+):(\\d+)-(\\d+):(\\d+)'
TYPE_RE     = r'(面交|冷凍宅配|宅配|駐店|活動|私定)'
```

### 2.3 樣本 dump（全 {len(in_person_c2)} 筆）

| # | 原始字串 | 地點 | 日期 | 類型 |
|---|---|---|---|---|""")
for i, r in enumerate(c2_analysis):
    md.append(f"| {i+1} | `{r['raw']}` | {r['location'] or '❌'} | {r['date'] or '❌'} | {r['type'] or '❌'} |")

md.append(f"""

---

## 3. 賣貨便 c12「⚠️必填！指定出貨日」樣本

- 共 **{len(seller_buy_c12_shipping)} 筆**訂單有指定出貨日
- regex 抓到：**{len(shipping_dates_extracted)} 筆**（{100*len(shipping_dates_extracted)/max(len(seller_buy_c12_shipping),1):.0f}%）
- regex 打不下：**{len(shipping_dates_failed)} 筆**

**regex**：`指定出貨日.*?(\\d+)/(\\d+)（(.)）`

### 3.1 各出貨日訂單分佈

| 出貨日 | 訂單數 |
|---|---|""")
for date, count in sorted(batch_dist.items()):
    md.append(f"| {date} | {count} |")

if shipping_dates_failed:
    md.append(f"\n### 3.2 regex 打不下樣本\n")
    for f in shipping_dates_failed:
        md.append(f"- `{f}`")

md.append(f"""

---

## 4. KOL 未完成 sheet 合成邊界驗證

### 4.1 合成規則
`c1（折扣碼）或 c2（IG 帳號）有值 = 新一筆 KOL 開始`

### 4.2 驗證結果
- 原始列數：954 列（含空白列）
- 合成後 KOL 數：**{len(kol_records)} 位**

### 4.3 KOL c4 寄貨時間值分佈（前 30）

| c4 值（截 30 字） | 出現次數 | 型態 |
|---|---|---|""")
for val, count in sorted(kol_c4_shipping.items(), key=lambda x: -x[1])[:30]:
    is_date = bool(re.match(r'^\d{4}-\d{2}-\d{2}', val))
    md.append(f"| `{val}` | {count} | {'📅 datetime' if is_date else '🔤 string'} |")

md.append(f"""

### 4.4 每位 KOL 品項數分佈

| 品項數 | KOL 人數 |
|---|---|""")
kol_dist = Counter(len(k["products"]) for k in kol_records)
for n, c in sorted(kol_dist.items()):
    md.append(f"| {n} 個 | {c} |")

md.append(f"""

---

## 5. 異常盤點

### 5.1 賣貨便 c22（配送數量）> 1

**共 {len(c22_over1)} 筆**，佔 {100*len(c22_over1)/max(len(seller_buy_orders),1):.1f}%。

⚠️ **待雇主 confirm**：c22 = 2 表示分兩箱寄——**出 2 張標籤還是 1 張**？

| 訂單編號 | c22 |
|---|---|""")
for x in c22_over1[:20]:
    md.append(f"| `{x['id']}` | {x['c22']} |")
if len(c22_over1) > 20:
    md.append(f"\n... (還有 {len(c22_over1)-20} 筆)")

md.append(f"""

### 5.2 賣貨便金額對帳（防護 #2）

**正確公式**：`預期 = Σ品項小計 + 運費 − 賣家折價券(c18) − 平台運費券(c19) − 平台折價券(c20)`

- ✅ 對得上：**{seller_buy_ok} 筆**
- ❌ 對不上：**{len(seller_buy_mismatches)} 筆**

{'✅ 100% 對得上，公式正確。' if not seller_buy_mismatches else '對不上的（前 20）：'}
""")
if seller_buy_mismatches:
    md.append("| 訂單編號 | 品項和 | 運費 | 折抵 | 預期 | 實際 c21 | 差額 |")
    md.append("|---|---|---|---|---|---|---|")
    for m in seller_buy_mismatches[:20]:
        md.append(f"| `{m['id']}` | {m['sub_sum']} | {m['freight']} | {m['discount']} | {m['expected']} | {m['total']} | {m['diff']} |")

_c28_date_pattern = re.compile(r'\d+/\d+|\d+月\d+')
c28_with_date = [n for n in seller_buy_c28_notes if _c28_date_pattern.search(n)]
md.append(f"""

### 5.3 賣家備註（c28）藏日期指示樣本

共 {len(seller_buy_c28_notes)} 筆非空備註，其中 **{len(c28_with_date)} 筆疑似含日期指示**：
""")
for n in c28_with_date[:10]:
    md.append(f"- `{n[:100]}`")

md.append(f"""

⚠️ 憲章要求：**若備註寫日期但 c12 選了不同日期，pipeline flag 而非猜**。

### 5.4 面交表非「面交」列

從 c2 樣本推斷有 {sum(1 for r in c2_analysis if r['type'] not in ('面交',))} 筆 c2 不是純面交。
- 冷凍宅配 → 歸「宅配」（雇主已確認）
- 駐店/活動 → 「待分類」桶（雇主已確認）

---

## 6. 待跟雇主 confirm 的事

1. **c22=2 出 2 張標籤還是 1 張**？（本批 {len(c22_over1)} 筆會受影響）
2. **未歸類品項字串**（見 §1.2）是否真為新品或誤植
3. **KOL 品項字串 alias** 對應是否正確（見 §1.4）
4. **標籤實體規格** — 熱感應標籤機 or A4 貼紙紙？
5. **賣貨便 c17 運費**（買家付的）算不算 NARCOS.sugar 營收
6. **賣家備註 c28 含日期指示** 優先度（vs c12）
7. **KOL 未完成/已完成 sheet** 進度狀態機
8. **賣貨便 c5 狀態**「訂單成立(尚未付款)」是否算進本批出爐量
9. **面交 c17「焙茶巴斯克 $980」vs 賣貨便「焙茶栗子巴斯克 $880」**是否為不同品項
10. **`長型4入_兩磅`**：display_name 尾巴被 Excel 截斷少「）」，是否用 alias 之一為權威名？

---

## 7. 下一步

- [x] skeleton 建好、fixtures 存好（本地、不進 git）
- [x] `docs/data-analysis.md` 產出（本檔）
- [x] `data/menu.yaml` v0 draft 產出（全中文 identifier）
- [ ] Yen review menu.yaml
- [ ] Yen confirm 「Claude Code + MCP」架構理解
- [ ] 寫正式 `docs/spec.md`（含資料模型、REST API、MCP tool、UI 頁面）
""")

with open(OUT_MD, "w", encoding="utf-8") as f:
    f.write("\n".join(md))

# ============================================================
# 7. menu.yaml 產出
# ============================================================
yaml_lines = [
    "# narcos-oven 官方菜單對照表 v0 (draft)",
    "# 生成自 fixtures/2026-07-round1/ 三個 xlsx。以賣貨便命名為權威。",
    "# 這是主軌 deterministic pipeline 的唯一 source of truth。",
    "# 識別碼慣例：atoms + products id 全中文（短、無空白/標點）",
    "# Yen 每次遇到新品項、由 Yen 更新此檔；AI 只做「建議」不做「決策」。",
    "",
    "# ---------- ATOMS ---------- ",
    "# 最小可拆解單位。出爐統計以 atom 為單位計數。",
    "atoms:",
]
for atom_id, meta in ATOMS.items():
    yaml_lines.append(f"  {atom_id}: {{unit: {meta['unit']}}}")
yaml_lines.append("")
yaml_lines.append("# ---------- PRODUCTS ---------- ")
yaml_lines.append("# 上架 SKU。display_name = 賣貨便權威命名；aliases = 別名（面交/KOL/舊稱）")
yaml_lines.append("products:")
for sku_id, meta in sorted(menu["products"].items()):
    yaml_lines.append(f"  {sku_id}:")
    yaml_lines.append(f"    display_name: \"{meta['display_name']}\"")
    yaml_lines.append(f"    category: {meta['category']}   # combo | single")
    aliases = list(dict.fromkeys(meta["aliases"]))
    if aliases:
        yaml_lines.append(f"    aliases:")
        for a in aliases:
            yaml_lines.append(f"      - \"{a}\"")
    else:
        yaml_lines.append(f"    aliases: []")
    yaml_lines.append(f"    contains:")
    for c in meta["contains"]:
        yaml_lines.append(f"      - {{atom: {c['atom']}, count: {c['count']}}}")
    yaml_lines.append(f"    price: {meta['price'] if meta['price'] is not None else 'null   # ⚠️ 待 Yen 補'}")
    yaml_lines.append(f"    cost: null   # 待 Yen 補（淨營收用）")
    yaml_lines.append(f"    seen_count: {meta['seen_count']}")
    yaml_lines.append("")

yaml_lines.append("# ---------- UNRECOGNIZED ---------- ")
yaml_lines.append("# 以下字串 auto-mapper 沒歸類到：")
if not unrecognized:
    yaml_lines.append("# ✅ 全部歸類完成")
for u in unrecognized:
    yaml_lines.append(f"# - \"{u['name']}\" (出現 {u['count']} 次)")

# ⚠️ v4 起不再直接覆寫 menu.yaml（Yen 手改為主）
# 改成寫到 docs/menu-proposal.yaml 讓 Yen review 後手動 merge
OUT_PROPOSAL = ROOT / "docs" / "menu-proposal.yaml"
with open(OUT_PROPOSAL, "w", encoding="utf-8") as f:
    f.write("# ⚠️ 這是 analyzer 從本批資料歸納的建議 menu 結構，僅供參考。\n")
    f.write("# 請 Yen 對照 data/menu.yaml 手動 merge、不要直接複製覆蓋。\n\n")
    f.write("\n".join(yaml_lines))

print(f"✅ {OUT_MD}")
print(f"✅ {OUT_PROPOSAL}（建議、不直接覆寫 data/menu.yaml）")
print()
print(f"統計（v3，全中文 identifier）：")
print(f"  賣貨便訂單: {len(seller_buy_orders)} 筆")
print(f"  賣貨便 unique 品項: {len(all_seller)} 個")
print(f"  面交 c2 有效列: {len(in_person_c2)} 筆")
print(f"  KOL 合作: {len(kol_records)} 位")
print(f"  出貨日 marker: {len(seller_buy_c12_shipping)} 筆（regex 抓到 {len(shipping_dates_extracted)}）")
print(f"  c22>1 訂單: {len(c22_over1)} 筆")
print(f"  金額對帳: {seller_buy_ok} ok / {len(seller_buy_mismatches)} 不符")
print(f"  未歸類品項: {len(unrecognized)} 個")
print(f"  atoms: {len(ATOMS)} 個")
print(f"  products: {len(menu['products'])} 個 SKU")
